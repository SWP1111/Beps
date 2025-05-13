import logging
import log_config
import os
from openpyxl import load_workbook
import pandas as pd
import time
import traceback
from config import Config
from services.statistics_excel_sheet_content import get_statistics_data, format_seconds_to_hhmmss
from services.statistics_excel_sheet_user import get_statistics_user_data

STYLE = """
<style>
  body {
    font-family: Arial, sans-serif;
    padding: 20px;
  }
  table {
    border-collapse: collapse;
    width: 100%;
    margin-bottom: 30px;
  }
  th, td {
    border: 1px solid #ccc;
    padding: 3px 4px;
    text-align: center;
  }
  th {
    background-color: #f5f5f5;
  }
  td {
    font-size: 12px;
  }
</style>
"""

def delete_old_files(folder_path, max_age_seconds=3600):
    """
    지정된 폴더에서 오래된 파일 삭제
    """
    if not os.path.exists(folder_path):
        logging.debug(f"Folder does not exist: {folder_path}")
        return

    now = time.time()
    for filename in os.listdir(folder_path):
        file_path = os.path.join(folder_path, filename)
        if os.path.isfile(file_path):
            file_age = now - os.path.getmtime(file_path)
            if file_age > max_age_seconds:
                os.remove(file_path)
                logging.debug(f"Deleted old file: {file_path}")

def scheduled_cleanup():
    """
    주기적으로 오래된 파일 삭제
    """
    try:
        delete_old_files(folder_path=Config.UPLOAD_DIR, max_age_seconds=3600)  # 1시간 이상된 파일 삭제
    except Exception as e:
        logging.error(f"Error during scheduled cleanup: {str(e)}, {traceback.format_exc()}")

def export_statistics_to_excel(path, filename, period_type, period_value, filter_type, filter_value):
    """
    통계 데이터를 엑셀로 내보내기
    """
    from services.user_summary_service import get_period_value
    
    start_date, end_date = get_period_value(period_type, period_value)
    
    files = get_statistics_data(start_date, end_date, filter_type, filter_value)  # 통계 데이터 가져오기
    rows = []

    if files:
        prev_top = prev_mid = None
        for f in files:
            top = f['top_name']
            mid = f['mid_name']
            bottom = f['bottom_name']
            avg_seconds = f['avg_stay_duration']
            avg_time = format_seconds_to_hhmmss(avg_seconds)
            
            if prev_top is not None and top != prev_top:
                rows.append({
                    '대분류': '',
                    '중분류': '',
                    '소분류': '',
                    '평균학습시간': '',
                    '의견서 수':'',
                    '최종 업데이트 날짜': '',
                    '관리자':'',
                })
                
            row = {
                '대분류': top if top != prev_top else '',
                '중분류': mid if mid != prev_mid else '',
                '소분류': bottom,
                '평균학습시간': avg_time,
                '의견서 수': f'{f['memo_count']}건',
                '최종 업데이트 날짜': f['update_at'],
                '관리자': f['manager_name'],
            }
            rows.append(row)
        
            prev_top = top
            prev_mid = mid
    
    df = pd.DataFrame(rows)
    
    usres = get_statistics_user_data(period_type, period_value, filter_type, filter_value)  # 사용자 통계 데이터 가져오기
    user_rows = []
    if usres:
        prev_company = prev_department = prev_name = None
        for u in usres:
            company = u['company']
            department = u['department']
            name = u['name']
            
            row = {
                '회사': '',
                '부서': '',
                '이름': name if name != prev_name else '',
                '총학습시간': '',
                '평균학습시간': '',
                '카테고리': u['category_name'],
                '학습시간': u['learning_time'],
                '의견서 수': f'{u['memo_count']}건',
            }
            if (company != prev_company) or (department != prev_department) or (name != prev_name):
                row['회사'] = company
                row['부서'] = department if department != '' else '-'
                row['이름'] = name if name != '' else '-'
                row['총학습시간'] = u['total_learning_time']
                row['평균학습시간'] = u['avg_learning_time']
            
            user_rows.append(row)
            prev_company = company
            prev_department = department
            prev_name = name
            
    df_user = pd.DataFrame(user_rows)
        
    excel_path = f"{path}/{filename}.xlsx"
    logging.debug(f"엑셀 파일 저장 경로: {excel_path}")
    os.makedirs(path, exist_ok=True)  # 디렉토리 생성
    
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='전체컨텐츠',index=False,startrow=2)
        df_user.to_excel(writer, sheet_name='회사&팀&직원',index=False,startrow=2)
   
    wb = load_workbook(excel_path)
    ws = wb['전체컨텐츠']
    ws.cell(row=1, column=3).value = f'{start_date} ~ {end_date}'
    wb.save(excel_path)
    wb.close()
    
    wb_user = load_workbook(excel_path)
    ws_user = wb_user['회사&팀&직원']
    ws_user.cell(row=1, column=1).value = '회사별,팀별,팀원별 기록(전체 확인 가능)'
    ws_user.cell(row=1, column=3).value = f'{start_date} ~ {end_date}'
    wb_user.save(excel_path)
    wb_user.close()
    
    df_content_html = generate_html_with_style(df) #df.to_html(index=False, classes='content_table', border=1)
    df_user_html = generate_html_with_style(df_user) #df_user.to_html(index=False, classes='user_table', border=1)
    
    content_html_path = os.path.join(path, f"{filename}_content.html")
    user_html_path = os.path.join(path, f"{filename}_user.html")
    
    with open(content_html_path, 'w', encoding='utf-8') as f:
        f.write(df_content_html)
    with open(user_html_path, 'w', encoding='utf-8') as f:
        f.write(df_user_html)
    
    return {
        'excel_path': excel_path,
        'html_content_name': f"{filename}_content.html",
        'html_user_name': f"{filename}_user.html"
    }
    

def generate_html_with_style(df):
    return f"""
    <!DOCTYPE html>
    <html lang="ko">
    <head>
      <meta charset="UTF-8">
      <title>미리보기</title>
      {STYLE}
    </head>
    <body>
      {df.to_html(index=False)}
    </body>
    </html>
    """              

    